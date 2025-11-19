from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse
from django.utils.html import strip_tags
from django.conf import settings
from django.template.loader import render_to_string
from .models import InformacionBasica, CalculoExperiencia, ExperienciaLaboral, InformacionAcademica, Posgrado, Especializacion, DocumentosIdentidad, Antecedentes, AnexosAdicionales
from .forms import (
    InformacionBasicaPublicForm,
    InformacionBasicaForm,
    ExperienciaLaboralFormSet,
    InformacionAcademicaFormSet,
    PosgradoFormSet,
    EspecializacionFormSet,
    DocumentosIdentidadForm,
    AntecedentesForm,
    AnexosAdicionalesForm,
)
import zipfile
import io
import os
import json
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import logging
import threading
import pytz

logger = logging.getLogger(__name__)

# Zona horaria de Colombia
COLOMBIA_TZ = pytz.timezone('America/Bogota')

# Gmail API scope
SCOPES = ['https://www.googleapis.com/auth/gmail.send']

def calcular_experiencia_total(informacion_basica):
    """Calcula automáticamente la experiencia total de una persona"""
    experiencias = informacion_basica.experiencias_laborales.all()

    total_meses = sum(exp.meses_experiencia for exp in experiencias)
    total_dias = sum(exp.dias_experiencia for exp in experiencias)

    # Convertir a años (considerando 12 meses por año)
    total_anos = round(total_meses / 12, 2)

    # Calcular años y meses para formato legible
    anos = total_meses // 12
    meses_restantes = total_meses % 12
    anos_y_meses = f"{anos} años y {meses_restantes} meses"

    # Crear o actualizar el registro de cálculo
    calculo, created = CalculoExperiencia.objects.update_or_create(
        informacion_basica=informacion_basica,
        defaults={
            'total_meses_experiencia': total_meses,
            'total_dias_experiencia': total_dias,
            'total_experiencia_anos': total_anos,
            'anos_y_meses_experiencia': anos_y_meses,
        }
    )
    return calculo

def enviar_correo_confirmacion(informacion_basica):
    """Envía correo de confirmación al usuario que completó el formulario usando Gmail API"""
    try:
        # Cargar credenciales: primero de variable de entorno (Railway) o archivo (desarrollo local)
        token_data = None
        
        # Intentar cargar desde variable de entorno (Railway)
        gmail_token_json = os.getenv('GMAIL_TOKEN_JSON')
        if gmail_token_json:
            try:
                token_data = json.loads(gmail_token_json)
                logger.info('Credenciales de Gmail cargadas desde variable de entorno')
            except json.JSONDecodeError as e:
                logger.error(f'Error parseando GMAIL_TOKEN_JSON: {str(e)}')
                return False
        
        # Si no hay variable de entorno, intentar cargar desde archivo (desarrollo local)
        if not token_data:
            BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            token_path = os.path.join(BASE_DIR, 'token.json')
            
            if os.path.exists(token_path):
                try:
                    with open(token_path, 'r') as token_file:
                        token_data = json.load(token_file)
                    logger.info(f'Credenciales de Gmail cargadas desde {token_path}')
                except Exception as e:
                    logger.error(f'Error leyendo token.json: {str(e)}')
                    return False
            else:
                logger.error(f'No se encontró token.json en {token_path} ni variable GMAIL_TOKEN_JSON')
                return False
        
        # Crear credenciales desde el token
        creds = Credentials(
            token=token_data.get('token'),
            refresh_token=token_data.get('refresh_token'),
            token_uri=token_data.get('token_uri'),
            client_id=token_data.get('client_id'),
            client_secret=token_data.get('client_secret'),
            scopes=token_data.get('scopes', SCOPES)
        )
        
        # Refrescar el token si es necesario
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            
        # Construir el servicio de Gmail
        service = build('gmail', 'v1', credentials=creds)
        
        # Preparar contexto para el template - usar zona horaria de Colombia
        fecha_colombia = datetime.now(COLOMBIA_TZ)
        context = {
            'nombre_completo': informacion_basica.nombre_completo,
            'cedula': informacion_basica.cedula,
            'correo': informacion_basica.correo,
            'telefono': informacion_basica.telefono,
            'fecha_registro': fecha_colombia.strftime('%d/%m/%Y %H:%M'),
        }

        # Renderizar template HTML
        html_message = render_to_string('formapp/email_confirmacion.html', context)

        # Crear mensaje de correo con nombre del remitente
        message = MIMEMultipart('alternative')
        message['To'] = informacion_basica.correo
        message['From'] = f'Sistema de Contratación CHVS <{settings.DEFAULT_FROM_EMAIL}>'
        message['Subject'] = 'Confirmación de Registro - Sistema de Contratación CHVS'
        
        # Adjuntar contenido HTML
        html_part = MIMEText(html_message, 'html', 'utf-8')
        message.attach(html_part)
        
        # Codificar el mensaje en base64
        raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
        send_message = {'raw': raw_message}
        
        # Enviar correo
        service.users().messages().send(userId='me', body=send_message).execute()
        
        logger.info(f'Correo enviado exitosamente a {informacion_basica.correo} vía Gmail API')
        return True
        
    except Exception as e:
        logger.error(f'Error al enviar correo a {informacion_basica.correo}: {str(e)}')
        return False

def public_form_view(request):
    if request.method == 'POST':
        form = InformacionBasicaPublicForm(request.POST, request.FILES)
        # Obtener género para pasarlo al formulario de documentos
        genero = request.POST.get('genero', '')
        documentos_form = DocumentosIdentidadForm(request.POST, request.FILES, genero=genero)
        antecedentes_form = AntecedentesForm(request.POST, request.FILES)
        anexos_form = AnexosAdicionalesForm(request.POST, request.FILES)
        experiencia_formset = ExperienciaLaboralFormSet(request.POST, request.FILES)
        academica_formset = InformacionAcademicaFormSet(request.POST, request.FILES)
        posgrado_formset = PosgradoFormSet(request.POST, request.FILES)
        especializacion_formset = EspecializacionFormSet(request.POST, request.FILES)

        # CRÍTICO: Validar TODOS los formularios ANTES de guardar cualquier cosa
        # Esto previene el bug donde certificados no se guardan pero el usuario ve "éxito"
        if form.is_valid():
            # Validar todos los formsets y formularios primero
            documentos_valid = documentos_form.is_valid()
            antecedentes_valid = antecedentes_form.is_valid()
            anexos_valid = anexos_form.is_valid()
            experiencia_valid = experiencia_formset.is_valid()
            academica_valid = academica_formset.is_valid()
            posgrado_valid = posgrado_formset.is_valid()
            especializacion_valid = especializacion_formset.is_valid()

            # Solo proceder si TODO es válido
            if documentos_valid and antecedentes_valid and anexos_valid and experiencia_valid and academica_valid and posgrado_valid and especializacion_valid:
                try:
                    with transaction.atomic():
                        informacion_basica = form.save()

                        # Guardar documentos de identidad
                        documentos = documentos_form.save(commit=False)
                        documentos.informacion_basica = informacion_basica
                        documentos.save()

                        # Guardar antecedentes
                        antecedentes = antecedentes_form.save(commit=False)
                        antecedentes.informacion_basica = informacion_basica
                        antecedentes.save()

                        # Guardar anexos adicionales (opcional)
                        anexos = anexos_form.save(commit=False)
                        anexos.informacion_basica = informacion_basica
                        anexos.save()

                        # Asociar formsets con la instancia guardada y guardar
                        experiencia_formset.instance = informacion_basica
                        experiencia_formset.save()
                        # Calcular experiencia automáticamente
                        calcular_experiencia_total(informacion_basica)

                        academica_formset.instance = informacion_basica
                        academica_formset.save()

                        posgrado_formset.instance = informacion_basica
                        posgrado_formset.save()

                        especializacion_formset.instance = informacion_basica
                        especializacion_formset.save()

                        # Enviar correo de confirmación al usuario en un thread separado
                        # para no bloquear la respuesta del formulario
                        def enviar_correo_async():
                            try:
                                enviar_correo_confirmacion(informacion_basica)
                            except Exception as e:
                                logger.error(f'Error en thread de correo: {str(e)}')

                        thread = threading.Thread(target=enviar_correo_async)
                        thread.daemon = True
                        thread.start()

                        messages.success(request, '¡Formulario enviado con éxito! Recibirás un correo de confirmación en los próximos minutos.')
                        return redirect('formapp:public_form')
                except Exception as e:
                    messages.error(request, f'Error al guardar el formulario: {str(e)}')
            else:
                # Mostrar errores específicos de cada formulario/formset que falló
                if not documentos_valid:
                    for field, error_list in documentos_form.errors.items():
                        for error in error_list:
                            messages.error(request, f'Error en Documentos de Identidad - {field}: {error}')

                if not antecedentes_valid:
                    for field, error_list in antecedentes_form.errors.items():
                        for error in error_list:
                            messages.error(request, f'Error en Antecedentes - {field}: {error}')

                if not experiencia_valid:
                    for i, form_errors in enumerate(experiencia_formset.errors):
                        if form_errors:
                            for field, error_list in form_errors.items():
                                for error in error_list:
                                    messages.error(request, f'Error en Experiencia Laboral #{i+1} - {field}: {error}')

                if not academica_valid:
                    for i, form_errors in enumerate(academica_formset.errors):
                        if form_errors:
                            for field, error_list in form_errors.items():
                                for error in error_list:
                                    messages.error(request, f'Error en Información Académica #{i+1} - {field}: {error}')

                if not posgrado_valid:
                    for i, form_errors in enumerate(posgrado_formset.errors):
                        if form_errors:
                            for field, error_list in form_errors.items():
                                for error in error_list:
                                    messages.error(request, f'Error en Posgrado #{i+1} - {field}: {error}')

                if not especializacion_valid:
                    for i, form_errors in enumerate(especializacion_formset.errors):
                        if form_errors:
                            for field, error_list in form_errors.items():
                                for error in error_list:
                                    messages.error(request, f'Error en Especialización #{i+1} - {field}: {error}')

                messages.warning(request, 'Por favor corrija los errores en el formulario antes de enviarlo.')
        else:
            messages.warning(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = InformacionBasicaPublicForm()
        documentos_form = DocumentosIdentidadForm()
        antecedentes_form = AntecedentesForm()
        anexos_form = AnexosAdicionalesForm()
        experiencia_formset = ExperienciaLaboralFormSet()
        academica_formset = InformacionAcademicaFormSet()
        posgrado_formset = PosgradoFormSet()
        especializacion_formset = EspecializacionFormSet()

    context = {
        'form': form,
        'documentos_form': documentos_form,
        'antecedentes_form': antecedentes_form,
        'anexos_form': anexos_form,
        'experiencia_formset': experiencia_formset,
        'academica_formset': academica_formset,
        'posgrado_formset': posgrado_formset,
        'especializacion_formset': especializacion_formset,
    }
    return render(request, 'formapp/public_form.html', context)

class ApplicantListView(LoginRequiredMixin, ListView):
    model = InformacionBasica
    template_name = 'formapp/applicant_list.html'
    context_object_name = 'applicants'
    ordering = ['-id']
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        # Búsqueda por cédula o nombre
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(cedula__icontains=search_query) |
                Q(nombre_completo__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Estadísticas correctamente calculadas
        context['total_personal'] = InformacionBasica.objects.count()
        context['con_experiencia'] = InformacionBasica.objects.filter(
            experiencias_laborales__isnull=False
        ).distinct().count()
        context['profesionales'] = InformacionBasica.objects.filter(
            formacion_academica__isnull=False
        ).distinct().count()
        context['con_posgrado'] = InformacionBasica.objects.filter(
            posgrados__isnull=False
        ).distinct().count()
        # Mantener el valor de búsqueda en el contexto
        context['search_query'] = self.request.GET.get('search', '')
        return context

class ApplicantDetailView(LoginRequiredMixin, DetailView):
    model = InformacionBasica
    template_name = 'formapp/applicant_detail.html'
    context_object_name = 'applicant'

@login_required
def applicant_edit_view(request, pk):
    """Vista para editar un registro existente"""
    applicant = get_object_or_404(InformacionBasica, pk=pk)

    # Obtener o crear instancias relacionadas
    try:
        documentos_identidad = applicant.documentos_identidad
    except DocumentosIdentidad.DoesNotExist:
        documentos_identidad = None

    try:
        antecedentes = applicant.antecedentes
    except Antecedentes.DoesNotExist:
        antecedentes = None

    try:
        anexos_adicionales = applicant.anexos_adicionales
    except AnexosAdicionales.DoesNotExist:
        anexos_adicionales = None

    if request.method == 'POST':
        form = InformacionBasicaForm(request.POST, request.FILES, instance=applicant)
        genero = request.POST.get('genero', applicant.genero)
        documentos_form = DocumentosIdentidadForm(request.POST, request.FILES, instance=documentos_identidad, genero=genero)
        antecedentes_form = AntecedentesForm(request.POST, request.FILES, instance=antecedentes)
        anexos_form = AnexosAdicionalesForm(request.POST, request.FILES, instance=anexos_adicionales)
        experiencia_formset = ExperienciaLaboralFormSet(request.POST, request.FILES, instance=applicant)
        academica_formset = InformacionAcademicaFormSet(request.POST, request.FILES, instance=applicant)
        posgrado_formset = PosgradoFormSet(request.POST, request.FILES, instance=applicant)
        especializacion_formset = EspecializacionFormSet(request.POST, request.FILES, instance=applicant)

        # Validar todos los formsets antes de guardar
        form_valid = form.is_valid()
        documentos_valid = documentos_form.is_valid()
        antecedentes_valid = antecedentes_form.is_valid()
        anexos_valid = anexos_form.is_valid()
        experiencia_valid = experiencia_formset.is_valid()
        academica_valid = academica_formset.is_valid()
        posgrado_valid = posgrado_formset.is_valid()
        especializacion_valid = especializacion_formset.is_valid()

        if form_valid and documentos_valid and antecedentes_valid and anexos_valid and experiencia_valid and academica_valid and posgrado_valid and especializacion_valid:
            try:
                with transaction.atomic():
                    informacion_basica = form.save()

                    # Guardar documentos de identidad
                    documentos = documentos_form.save(commit=False)
                    documentos.informacion_basica = informacion_basica
                    documentos.save()

                    # Guardar antecedentes
                    antecedentes_obj = antecedentes_form.save(commit=False)
                    antecedentes_obj.informacion_basica = informacion_basica
                    antecedentes_obj.save()

                    # Guardar anexos adicionales
                    anexos = anexos_form.save(commit=False)
                    anexos.informacion_basica = informacion_basica
                    anexos.save()

                    # Guardar todos los formsets
                    # Guardar experiencia laboral - usar save() que maneja automáticamente todo
                    experiencia_formset.save()
                    
                    # Recalcular meses y días para cada experiencia guardada
                    for experiencia in informacion_basica.experiencias_laborales.all():
                        if experiencia.fecha_inicial and experiencia.fecha_terminacion:
                            from datetime import datetime as dt
                            fecha_inicio = dt.combine(experiencia.fecha_inicial, dt.min.time())
                            fecha_fin = dt.combine(experiencia.fecha_terminacion, dt.min.time())
                            
                            # Calcular diferencia
                            delta = fecha_fin - fecha_inicio
                            total_dias = delta.days
                            
                            # Calcular meses
                            anos = fecha_fin.year - fecha_inicio.year
                            meses = fecha_fin.month - fecha_inicio.month
                            dias = fecha_fin.day - fecha_inicio.day
                            
                            if dias < 0:
                                meses -= 1
                                # Obtener días del mes anterior
                                if fecha_inicio.month == 1:
                                    ultimo_dia = dt(fecha_inicio.year - 1, 12, 31).day
                                else:
                                    ultimo_dia = dt(fecha_inicio.year, fecha_inicio.month - 1, 1).day
                                dias += ultimo_dia
                            
                            if meses < 0:
                                anos -= 1
                                meses += 12
                            
                            total_meses = (anos * 12) + meses
                            
                            experiencia.meses_experiencia = total_meses
                            experiencia.dias_experiencia = total_dias
                            experiencia.save()
                    
                    # Calcular experiencia total automáticamente
                    calcular_experiencia_total(informacion_basica)

                    # Guardar los demás formsets
                    academica_formset.save()
                    posgrado_formset.save()
                    especializacion_formset.save()

                    messages.success(request, f'Registro de {informacion_basica.nombre_completo} actualizado con éxito!')
                    return redirect('formapp:applicant_detail', pk=informacion_basica.pk)
            except Exception as e:
                import traceback
                messages.error(request, f'Error al actualizar el registro: {str(e)}')
                logger.error(f'Error guardando experiencia laboral: {str(e)}\n{traceback.format_exc()}')
        else:
            # Si hay errores, mostrar mensajes específicos
            if not form_valid:
                messages.error(request, 'Por favor corrija los errores en el formulario principal.')
            if not experiencia_valid:
                error_msg = experiencia_formset.non_form_errors()
                if error_msg:
                    messages.error(request, f'Error en Experiencia Laboral: {error_msg}')
                for idx, form_exp in enumerate(experiencia_formset, 1):
                    if form_exp.errors:
                        messages.error(request, f'Experiencia {idx}: {form_exp.errors}')
            if not academica_valid:
                messages.error(request, 'Por favor corrija los errores en Formación Académica.')
            if not posgrado_valid:
                messages.error(request, 'Por favor corrija los errores en Posgrados.')
            if not especializacion_valid:
                messages.error(request, 'Por favor corrija los errores en Especializaciones.')
    else:
        form = InformacionBasicaForm(instance=applicant)
        documentos_form = DocumentosIdentidadForm(instance=documentos_identidad, genero=applicant.genero)
        antecedentes_form = AntecedentesForm(instance=antecedentes)
        anexos_form = AnexosAdicionalesForm(instance=anexos_adicionales)
        experiencia_formset = ExperienciaLaboralFormSet(instance=applicant)
        academica_formset = InformacionAcademicaFormSet(instance=applicant)
        posgrado_formset = PosgradoFormSet(instance=applicant)
        especializacion_formset = EspecializacionFormSet(instance=applicant)

    context = {
        'form': form,
        'documentos_form': documentos_form,
        'antecedentes_form': antecedentes_form,
        'anexos_form': anexos_form,
        'experiencia_formset': experiencia_formset,
        'academica_formset': academica_formset,
        'posgrado_formset': posgrado_formset,
        'especializacion_formset': especializacion_formset,
        'applicant': applicant,
    }
    return render(request, 'formapp/applicant_edit.html', context)

@login_required
def applicant_delete_view(request, pk):
    """Vista para eliminar un registro"""
    applicant = get_object_or_404(InformacionBasica, pk=pk)

    if request.method == 'POST':
        nombre = applicant.nombre_completo
        try:
            applicant.delete()
            messages.success(request, f'Registro de {nombre} eliminado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar el registro: {str(e)}')
        return redirect('formapp:applicant_list')

    # Si no es POST, redirigir a la lista
    return redirect('formapp:applicant_list')

def create_excel_for_person(applicant):
    """Crea un archivo Excel con toda la información de una persona"""
    wb = Workbook()

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="2C3E50")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Hoja 1: Información Básica
    ws1 = wb.active
    ws1.title = "Información Básica"

    # Título
    ws1['A1'] = f"INFORMACIÓN PERSONAL - {applicant.nombre_completo}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')

    # Datos personales
    row = 3
    personal_data = [
        ("Cédula", applicant.cedula),
        ("Nombre Completo", applicant.nombre_completo),
        ("Género", applicant.genero),
        ("Dirección", f"{applicant.tipo_via} {applicant.numero_via} #{applicant.numero_casa}"),
        ("Complemento Dirección", applicant.complemento_direccion or "N/A"),
        ("Barrio", applicant.barrio or "N/A"),
        ("Teléfono", applicant.telefono),
        ("Correo", applicant.correo),
    ]

    for label, value in personal_data:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'A{row}'].fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        ws1[f'B{row}'] = value
        ws1[f'A{row}'].border = border
        ws1[f'B{row}'].border = border
        row += 1

    # Información profesional
    row += 2
    ws1[f'A{row}'] = "INFORMACIÓN PROFESIONAL"
    ws1[f'A{row}'].font = title_font
    ws1.merge_cells(f'A{row}:B{row}')
    row += 2

    professional_data = [
        ("Perfil", applicant.perfil or "N/A"),
        ("Área de Conocimiento", applicant.area_conocimiento or "N/A"),
        ("Área del Conocimiento", applicant.area_del_conocimiento or "N/A"),
        ("Tipo de Perfil", applicant.tipo_perfil or "N/A"),
        ("Profesión", applicant.profesion or "N/A"),
        ("Experiencia", applicant.experiencia or "N/A"),
        ("Tiempo de Experiencia", applicant.tiempo_experiencia or "N/A"),
        ("Cantidad", applicant.cantidad or "N/A"),
        ("Organización", applicant.organizacion or "N/A"),
        ("Contrato", applicant.contrato or "N/A"),
        ("Observaciones", applicant.observacion or "N/A"),
    ]

    for label, value in professional_data:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'A{row}'].fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        ws1[f'B{row}'] = value
        ws1[f'A{row}'].border = border
        ws1[f'B{row}'].border = border
        row += 1

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 50

    # Hoja 2: Experiencia Laboral
    ws2 = wb.create_sheet("Experiencia Laboral")
    ws2['A1'] = f"EXPERIENCIA LABORAL - {applicant.nombre_completo}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:H1')

    # Encabezados
    headers = ["Cargo", "Cargo Anexo 11", "Fecha Inicial", "Fecha Terminación",
               "Meses", "Días", "Objeto Contractual", "Funciones"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row = 4
    for exp in applicant.experiencias_laborales.all():
        ws2.cell(row=row, column=1, value=exp.cargo).border = border
        ws2.cell(row=row, column=2, value=exp.cargo_anexo_11).border = border
        ws2.cell(row=row, column=3, value=exp.fecha_inicial.strftime('%Y-%m-%d')).border = border
        ws2.cell(row=row, column=4, value=exp.fecha_terminacion.strftime('%Y-%m-%d')).border = border
        ws2.cell(row=row, column=5, value=exp.meses_experiencia).border = border
        ws2.cell(row=row, column=6, value=exp.dias_experiencia).border = border
        ws2.cell(row=row, column=7, value=exp.objeto_contractual).border = border
        ws2.cell(row=row, column=8, value=exp.funciones).border = border
        row += 1

    for col in range(1, 9):
        ws2.column_dimensions[chr(64 + col)].width = 20

    # Hoja 3: Información Académica
    ws3 = wb.create_sheet("Información Académica")
    ws3['A1'] = f"INFORMACIÓN ACADÉMICA - {applicant.nombre_completo}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:G1')

    # Encabezados
    headers = ["Profesión", "Universidad", "Tarjeta Profesional",
               "N° Tarjeta/Resolución", "Fecha Expedición", "Fecha Grado", "Meses Experiencia"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row = 4
    for academica in applicant.formacion_academica.all():
        ws3.cell(row=row, column=1, value=academica.profesion).border = border
        ws3.cell(row=row, column=2, value=academica.universidad).border = border
        ws3.cell(row=row, column=3, value=academica.tarjeta_profesional).border = border
        ws3.cell(row=row, column=4, value=academica.numero_tarjeta_resolucion or "N/A").border = border
        ws3.cell(row=row, column=5, value=academica.fecha_expedicion.strftime('%Y-%m-%d') if academica.fecha_expedicion else "N/A").border = border
        ws3.cell(row=row, column=6, value=academica.fecha_grado.strftime('%Y-%m-%d')).border = border
        ws3.cell(row=row, column=7, value=academica.meses_experiencia_profesion).border = border
        row += 1

    for col in range(1, 8):
        ws3.column_dimensions[chr(64 + col)].width = 20

    # Hoja 4: Posgrados
    ws4 = wb.create_sheet("Posgrados")
    ws4['A1'] = f"POSGRADOS - {applicant.nombre_completo}"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:D1')

    # Encabezados
    headers = ["Nombre Posgrado", "Universidad", "Fecha Terminación", "Meses Experiencia"]
    for col, header in enumerate(headers, start=1):
        cell = ws4.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row = 4
    for posgrado in applicant.posgrados.all():
        ws4.cell(row=row, column=1, value=posgrado.nombre_posgrado).border = border
        ws4.cell(row=row, column=2, value=posgrado.universidad).border = border
        ws4.cell(row=row, column=3, value=posgrado.fecha_terminacion.strftime('%Y-%m-%d')).border = border
        ws4.cell(row=row, column=4, value=posgrado.meses_de_experiencia).border = border
        row += 1

    for col in range(1, 5):
        ws4.column_dimensions[chr(64 + col)].width = 25

    # Hoja 5: Especializaciones
    ws5 = wb.create_sheet("Especializaciones")
    ws5['A1'] = f"ESPECIALIZACIONES - {applicant.nombre_completo}"
    ws5['A1'].font = title_font
    ws5.merge_cells('A1:D1')

    # Encabezados
    headers = ["Nombre Especialización", "Universidad", "Fecha Terminación", "Meses Experiencia"]
    for col, header in enumerate(headers, start=1):
        cell = ws5.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row = 4
    for especializacion in applicant.especializaciones.all():
        ws5.cell(row=row, column=1, value=especializacion.nombre_especializacion).border = border
        ws5.cell(row=row, column=2, value=especializacion.universidad).border = border
        ws5.cell(row=row, column=3, value=especializacion.fecha_terminacion.strftime('%Y-%m-%d')).border = border
        ws5.cell(row=row, column=4, value=especializacion.meses_de_experiencia).border = border
        row += 1

    for col in range(1, 5):
        ws5.column_dimensions[chr(64 + col)].width = 25

    # Hoja 6: Cálculo de Experiencia
    ws6 = wb.create_sheet("Cálculo Experiencia")
    ws6['A1'] = f"CÁLCULO DE EXPERIENCIA - {applicant.nombre_completo}"
    ws6['A1'].font = title_font
    ws6.merge_cells('A1:B1')

    row = 3
    try:
        calculo = applicant.calculo_experiencia
        calculo_data = [
            ("Total Meses Experiencia", calculo.total_meses_experiencia),
            ("Total Días Experiencia", calculo.total_dias_experiencia),
            ("Total Experiencia (Años)", calculo.total_experiencia_anos),
            ("Años y Meses", calculo.anos_y_meses_experiencia),
        ]
    except:
        calculo_data = [
            ("Total Meses Experiencia", "No calculado"),
            ("Total Días Experiencia", "No calculado"),
            ("Total Experiencia (Años)", "No calculado"),
            ("Años y Meses", "No calculado"),
        ]

    for label, value in calculo_data:
        ws6[f'A{row}'] = label
        ws6[f'A{row}'].font = Font(bold=True)
        ws6[f'A{row}'].fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        ws6[f'B{row}'] = value
        ws6[f'A{row}'].border = border
        ws6[f'B{row}'].border = border
        row += 1

    ws6.column_dimensions['A'].width = 30
    ws6.column_dimensions['B'].width = 30

    return wb

def generar_anexo11_pdf(applicant):
    """
    Genera un PDF en formato ANEXO 11 con la información del candidato
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT

    # Función auxiliar para convertir números a texto en español
    def numero_a_texto_es(n):
        """Convierte números del 1 al 31 a texto en español"""
        numeros = {
            1: 'uno', 2: 'dos', 3: 'tres', 4: 'cuatro', 5: 'cinco',
            6: 'seis', 7: 'siete', 8: 'ocho', 9: 'nueve', 10: 'diez',
            11: 'once', 12: 'doce', 13: 'trece', 14: 'catorce', 15: 'quince',
            16: 'dieciséis', 17: 'diecisiete', 18: 'dieciocho', 19: 'diecinueve', 20: 'veinte',
            21: 'veintiuno', 22: 'veintidós', 23: 'veintitrés', 24: 'veinticuatro', 25: 'veinticinco',
            26: 'veintiséis', 27: 'veintisiete', 28: 'veintiocho', 29: 'veintinueve', 30: 'treinta',
            31: 'treinta y uno'
        }
        return numeros.get(n, str(n))

    # Diccionario de meses en español
    meses_es = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
        7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }

    # Crear buffer en memoria para el PDF
    pdf_buffer = io.BytesIO()

    # Crear documento
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )

    # Estilos
    styles = getSampleStyleSheet()

    # Estilo para el título
    titulo_style = ParagraphStyle(
        'TituloAnexo',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    # Estilo para subtítulos
    subtitulo_style = ParagraphStyle(
        'Subtitulo',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    # Estilo para texto normal
    normal_style = ParagraphStyle(
        'NormalText',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_JUSTIFY,
        spaceAfter=12
    )

    # Estilo para texto pequeño
    small_style = ParagraphStyle(
        'SmallText',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_LEFT
    )

    # Estilo para texto en celdas de tabla (con word wrap)
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_LEFT,
        leading=10,  # Espaciado entre líneas
        wordWrap='CJK'  # Permite ajuste de texto
    )

    # Estilo para texto centrado en celdas de tabla
    cell_center_style = ParagraphStyle(
        'CellCenterText',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        leading=10,
        wordWrap='CJK'
    )

    # Contenido del documento
    elementos = []

    # Título principal
    elementos.append(Paragraph("ANEXO 11", titulo_style))
    elementos.append(Paragraph("CARTA DE COMPROMISO PERSONAL", subtitulo_style))
    elementos.append(Spacer(1, 0.3*inch))

    # ==================== PÁGINA 1 ====================

    # Fecha y destinatario
    fecha_obj = datetime.now()
    dia = fecha_obj.day
    mes = fecha_obj.month
    anio = fecha_obj.year

    # Nombre del día en texto y mes en español
    dia_texto = numero_a_texto_es(dia)
    mes_nombre = meses_es.get(mes, 'error')

    # Fecha en formato "04 de noviembre de 2025"
    fecha_actual = f"{dia:02d} de {mes_nombre} de {anio}"

    elementos.append(Paragraph(f"Cali, {fecha_actual}", normal_style))
    elementos.append(Spacer(1, 0.2*inch))

    elementos.append(Paragraph("Señores:", normal_style))
    elementos.append(Paragraph("<b>SECRETARÍA DE BIENESTAR SOCIAL</b>", normal_style))
    elementos.append(Paragraph("<b>DISTRITO ESPECIAL DE SANTIAGO DE CALI</b>", normal_style))
    elementos.append(Paragraph("Ciudad", normal_style))
    elementos.append(Spacer(1, 0.2*inch))

    # Referencia - usar el campo contrato
    numero_proceso = applicant.contrato or "4146.010.32.1.2366.2025"
    elementos.append(Paragraph(f"<b>REFERENCIA:</b> Proceso No. {numero_proceso}", normal_style))
    elementos.append(Spacer(1, 0.2*inch))

    # Cuerpo de la carta - usar el campo organizacion y perfil
    organizacion = applicant.organizacion or "UNIÓN TEMPORAL COMISIÓN ARQUIDIOCESANA VIDA JUSTICIA Y PAZ 25-2"
    cargo_propuesto = applicant.perfil or "el cargo correspondiente"

    texto_compromiso = f"""
    Yo, <b>{applicant.nombre_completo}</b>, identificado con c.c. <b>{applicant.cedula}</b>,
    acepto ser presentado por la empresa <b>{organizacion}</b>
    como <b>{cargo_propuesto}</b> en su propuesta dentro de los equipo de profesionales, y participar dentro
    de la ejecución del proceso de selección No. <b>{numero_proceso}</b>, que tiene como objeto:
    AUNAR ESFUERZOS TÉCNICOS, HUMANOS, ADMINISTRATIVOS Y FINANCIEROS PARA EL MEJORAMIENTO DE LAS
    CONDICIONES DE SEGURIDAD ALIMENTARIA DE LA POBLACIÓN VULNERABLE, GARANTIZANDO SU ACCESO A LOS
    ALIMENTOS Y BRINDANDO INTERVENCIÓN PSICOSOCIAL, EN EL DISTRITO DE SANTIAGO DE CALI, DE CONFORMIDAD
    CON EL PROYECTO DE INVERSIÓN "FORTALECIMIENTO DEL PROGRAMA DE SEGURIDAD ALIMENTARIA Y NUTRICIONAL
    EN SANTIAGO DE CALI" - BP-26005417 de acuerdo con lo establecido en la invitación, el estudio
    previo y documento denominado ANEXO TÉCNICO.
    """

    elementos.append(Paragraph(texto_compromiso, normal_style))
    elementos.append(Spacer(1, 0.15*inch))

    elementos.append(Paragraph(
        "Por lo que me comprometo a formar parte del equipo de trabajo durante el plazo que dure el convenio de asociación.",
        normal_style
    ))
    elementos.append(Spacer(1, 0.2*inch))

    # Texto de firma con fecha dinámica en español
    texto_firma = f"Para constancia se firma a los {dia_texto} ({dia}) días del mes de {mes_nombre} del {anio}."
    elementos.append(Paragraph(texto_firma, normal_style))
    elementos.append(Spacer(1, 0.5*inch))

    # Tabla de firmas
    firmas_data = [
        ['_________________________', '_________________________'],
        [f'{applicant.nombre_completo}', 'Diego Fernando Guzmán Ruiz'],
        ['Firma del Profesional', 'Firma del Representante Legal']
    ]

    tabla_firmas = Table(firmas_data, colWidths=[3.5*inch, 3.5*inch])
    tabla_firmas.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 2), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, 0), 0),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
    ]))

    elementos.append(tabla_firmas)

    # ==================== SALTO DE PÁGINA ====================
    elementos.append(PageBreak())

    # ==================== PÁGINA 2 ====================

    # 1. Tabla: RELACIÓN DE EXPERIENCIA PROFESIONALES PARA EL PERSONAL BASE
    # Construir dirección completa
    direccion_completa = f"{applicant.tipo_via} {applicant.numero_via} #{applicant.numero_casa}"
    if applicant.complemento_direccion:
        direccion_completa += f" {applicant.complemento_direccion}"
    if applicant.barrio:
        direccion_completa += f", Barrio {applicant.barrio}"

    # Datos de la tabla con título en la primera fila con fondo gris
    # Usar Paragraph para textos largos que puedan desbordarse
    tabla_experiencia_data = [
        ['RELACIÓN DE EXPERIENCIA PROFESIONALES PARA EL PERSONAL BASE', ''],  # Título con span
        ['CARGO PROPUESTO:', Paragraph(str(cargo_propuesto or ''), cell_style)],
        ['NOMBRES Y APELLIDOS:', Paragraph(str(applicant.nombre_completo or ''), cell_style)],
        ['TIPO Y Nº DOCUMENTO DE IDENTIDAD:', Paragraph(f'CC {applicant.cedula}', cell_style)],
        ['DIRECCIÓN:', Paragraph(str(direccion_completa or ''), cell_style)],
        ['TELÉFONO:', Paragraph(str(applicant.telefono or ''), cell_style)],
        ['CORREO ELECTRÓNICO:', Paragraph(str(applicant.correo or ''), cell_style)],
    ]

    tabla_experiencia = Table(tabla_experiencia_data, colWidths=[2.5*inch, 4.5*inch])
    tabla_experiencia.setStyle(TableStyle([
        # Título - Primera fila con fondo gris
        ('SPAN', (0, 0), (1, 0)),  # Combinar columnas para el título
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#D3D3D3')),  # Gris
        ('TEXTCOLOR', (0, 0), (1, 0), colors.black),
        ('ALIGN', (0, 0), (1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (1, 0), 10),
        ('TOPPADDING', (0, 0), (1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (1, 0), 8),

        # Resto de las filas
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E8E8E8')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Cambiar a TOP para textos largos
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))

    elementos.append(tabla_experiencia)
    elementos.append(Spacer(1, 0.3*inch))

    # 2. Tabla: ESTUDIOS REALIZADOS - Formato de 4 columnas
    # Obtener todos los estudios
    formaciones_academicas = list(applicant.formacion_academica.all())
    posgrados = list(applicant.posgrados.all())
    especializaciones = list(applicant.especializaciones.all())

    # Calcular experiencia en años
    try:
        calculo_exp = applicant.calculo_experiencia
        experiencia_anos = f"{calculo_exp.total_experiencia_anos} años"
    except:
        experiencia_anos = "No calculada"

    # Construir contenido consolidado para cada columna
    # UNIVERSITARIOS - Agrupar todos los estudios universitarios
    contenido_titulos_univ = ''
    contenido_instituciones_univ = ''
    contenido_fechas_univ = ''
    tarjeta_texto = ''
    
    for formacion in formaciones_academicas:
        titulo = formacion.profesion or ''
        institucion = formacion.universidad or ''
        fecha = formacion.fecha_grado.strftime('%d/%m/%Y') if formacion.fecha_grado else ''
        
        if titulo:
            contenido_titulos_univ += f'{titulo}<br/>'
        if institucion:
            contenido_instituciones_univ += f'{institucion}<br/>'
        if fecha:
            contenido_fechas_univ += f'{fecha}<br/>'
        
        # Tarjeta profesional (solo la primera o consolidar todas)
        if not tarjeta_texto:
            if formacion.tarjeta_profesional == 'Tarjeta Profesional':
                tarjeta_texto = f"Tarjeta Profesional: {formacion.numero_tarjeta_resolucion or 'N/A'}"
            elif formacion.tarjeta_profesional == 'Resolución':
                tarjeta_texto = f"Resolución: {formacion.numero_tarjeta_resolucion or 'N/A'}"
    
    # ESPECIALIZACIÓN - Agrupar todas las especializaciones
    contenido_titulos_esp = ''
    contenido_instituciones_esp = ''
    contenido_fechas_esp = ''
    
    for especializacion in especializaciones:
        titulo = especializacion.nombre_especializacion or ''
        institucion = especializacion.universidad or ''
        fecha = especializacion.fecha_terminacion.strftime('%d/%m/%Y') if especializacion.fecha_terminacion else ''
        
        if titulo:
            contenido_titulos_esp += f'{titulo}<br/>'
        if institucion:
            contenido_instituciones_esp += f'{institucion}<br/>'
        if fecha:
            contenido_fechas_esp += f'{fecha}<br/>'
    
    # OTROS (POSGRADOS) - Agrupar todos los posgrados
    contenido_titulos_otros = ''
    contenido_instituciones_otros = ''
    contenido_fechas_otros = ''
    
    for posgrado in posgrados:
        titulo = posgrado.nombre_posgrado or ''
        institucion = posgrado.universidad or ''
        fecha = posgrado.fecha_terminacion.strftime('%d/%m/%Y') if posgrado.fecha_terminacion else ''
        
        if titulo:
            contenido_titulos_otros += f'{titulo}<br/>'
        if institucion:
            contenido_instituciones_otros += f'{institucion}<br/>'
        if fecha:
            contenido_fechas_otros += f'{fecha}<br/>'
    
    # Si no hay tarjeta profesional, poner "No Aplica"
    if not tarjeta_texto:
        tarjeta_texto = "No Aplica"

    # Construir la tabla con datos consolidados
    estudios_nueva_data = [
        # Fila de título con fondo gris
        ['ESTUDIOS REALIZADOS', '', '', ''],
        # Fila de encabezados de columnas
        ['DESCRIPCIÓN', 'UNIVERSITARIOS', 'ESPECIALIZACIÓN', 'OTROS'],
        # Fila de TÍTULO OBTENIDO con todos los títulos consolidados
        ['TÍTULO OBTENIDO',
         Paragraph(contenido_titulos_univ, cell_center_style) if contenido_titulos_univ else '',
         Paragraph(contenido_titulos_esp, cell_center_style) if contenido_titulos_esp else '',
         Paragraph(contenido_titulos_otros, cell_center_style) if contenido_titulos_otros else ''],
        # Fila de INSTITUCIÓN con todas las instituciones consolidadas
        ['INSTITUCIÓN',
         Paragraph(contenido_instituciones_univ, cell_center_style) if contenido_instituciones_univ else '',
         Paragraph(contenido_instituciones_esp, cell_center_style) if contenido_instituciones_esp else '',
         Paragraph(contenido_instituciones_otros, cell_center_style) if contenido_instituciones_otros else ''],
        # Fila de FECHA DE GRADO con todas las fechas consolidadas
        ['FECHA DE GRADO',
         Paragraph(contenido_fechas_univ, cell_center_style) if contenido_fechas_univ else '',
         Paragraph(contenido_fechas_esp, cell_center_style) if contenido_fechas_esp else '',
         Paragraph(contenido_fechas_otros, cell_center_style) if contenido_fechas_otros else ''],
    ]

    # Agregar fila de tarjeta profesional (solo para universitarios)
    if formaciones_academicas:
        estudios_nueva_data.append(['TARJETA PROFESIONAL', Paragraph(str(tarjeta_texto or ''), cell_center_style), '', ''])
    
    # Agregar fila de experiencia
    estudios_nueva_data.append(['2. EXPERIENCIA:', Paragraph(str(experiencia_anos or ''), cell_center_style), '', ''])

    tabla_estudios_nueva = Table(estudios_nueva_data, colWidths=[1.75*inch, 1.75*inch, 1.75*inch, 1.75*inch])
    
    # Calcular número de filas para los estilos
    num_filas = len(estudios_nueva_data)
    
    # Construir lista de estilos base
    estilos_base = [
        # Título - Primera fila con fondo gris y span
        ('SPAN', (0, 0), (3, 0)),
        ('BACKGROUND', (0, 0), (3, 0), colors.HexColor('#D3D3D3')),
        ('TEXTCOLOR', (0, 0), (3, 0), colors.black),
        ('ALIGN', (0, 0), (3, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (3, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (3, 0), 10),

        # Encabezados de columnas - Segunda fila
        ('BACKGROUND', (0, 1), (3, 1), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 1), (3, 1), colors.whitesmoke),
        ('ALIGN', (0, 1), (3, 1), 'CENTER'),
        ('FONTNAME', (0, 1), (3, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (3, 1), 9),

        # Primera columna (DESCRIPCIÓN) - Negrita para todas las filas de datos
        ('BACKGROUND', (0, 2), (0, num_filas - 1), colors.HexColor('#E8E8E8')),
        ('FONTNAME', (0, 2), (0, num_filas - 1), 'Helvetica-Bold'),
        ('ALIGN', (0, 2), (0, num_filas - 1), 'LEFT'),
        ('FONTSIZE', (0, 2), (0, num_filas - 1), 8),

        # Resto de datos
        ('FONTNAME', (1, 2), (3, num_filas - 1), 'Helvetica'),
        ('FONTSIZE', (1, 2), (3, num_filas - 1), 8),
        ('ALIGN', (1, 2), (3, num_filas - 1), 'CENTER'),

        # Bordes y espaciado
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # TOP para textos largos
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]
    
    # Agregar estilos para tarjeta profesional y experiencia (span solo en la primera columna)
    # Tarjeta profesional (si existe, es la penúltima fila)
    if formaciones_academicas:
        fila_tarjeta = num_filas - 2  # Penúltima fila
        estilos_base.append(('SPAN', (1, fila_tarjeta), (3, fila_tarjeta)))
    
    # Experiencia (última fila)
    fila_experiencia = num_filas - 1
    estilos_base.append(('SPAN', (1, fila_experiencia), (3, fila_experiencia)))
    
    tabla_estudios_nueva.setStyle(TableStyle(estilos_base))

    elementos.append(tabla_estudios_nueva)

    # Construir PDF
    doc.build(elementos)

    # Retornar buffer
    pdf_buffer.seek(0)
    return pdf_buffer

@login_required
def download_individual_zip(request, pk):
    """Descarga un ZIP con todos los certificados y Excel de una persona"""
    applicant = get_object_or_404(InformacionBasica, pk=pk)

    # Crear archivo ZIP en memoria
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # 1. Agregar Excel con toda la información
        wb = create_excel_for_person(applicant)
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        filename_safe = applicant.nombre_completo.replace(' ', '_')
        zip_file.writestr(
            f"{filename_safe}_Informacion.xlsx",
            excel_buffer.getvalue()
        )

        # 2. Agregar PDF ANEXO 11
        try:
            pdf_buffer = generar_anexo11_pdf(applicant)
            zip_file.writestr(
                f"{filename_safe}_ANEXO_11.pdf",
                pdf_buffer.getvalue()
            )
        except Exception as e:
            logger.error(f"Error al generar PDF ANEXO 11: {str(e)}")

        # 3. Agregar certificados laborales
        for idx, experiencia in enumerate(applicant.experiencias_laborales.all(), start=1):
            if experiencia.certificado_laboral:
                try:
                    # Leer el archivo del certificado usando context manager
                    certificado_file = experiencia.certificado_laboral
                    with certificado_file.open('rb') as f:
                        file_content = f.read()

                    # Obtener la extensión del archivo
                    # Cloudinary puede no incluir extensión en el nombre, así que intentamos obtenerla de la URL
                    ext = os.path.splitext(certificado_file.name)[1]
                    if not ext and hasattr(certificado_file, 'url'):
                        # Intentar obtener extensión de la URL de Cloudinary
                        url = certificado_file.url
                        # La URL de Cloudinary tiene formato: .../upload/v123456/archivo.ext
                        if '.' in url.split('/')[-1]:
                            ext = '.' + url.split('/')[-1].split('.')[-1].split('?')[0]

                    # Si aún no hay extensión, detectar por contenido
                    if not ext:
                        # Detectar tipo por magic bytes
                        if file_content.startswith(b'%PDF'):
                            ext = '.pdf'
                        elif file_content.startswith(b'\x89PNG'):
                            ext = '.png'
                        elif file_content.startswith(b'\xff\xd8\xff'):
                            ext = '.jpg'
                        else:
                            ext = '.pdf'  # Default a PDF si no se puede detectar

                    cargo_safe = experiencia.cargo.replace(' ', '_').replace('/', '-')

                    # Agregar al ZIP
                    zip_file.writestr(
                        f"Certificados/{idx}_{cargo_safe}{ext}",
                        file_content
                    )
                except Exception as e:
                    logger.error(f"Error al agregar certificado {idx} de {applicant.nombre_completo}: {e}")

    # Preparar respuesta
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename_safe}_Completo.zip"'

    return response

@login_required
def download_all_zip(request):
    """Descarga un ZIP con toda la información de TODO el personal"""
    applicants = InformacionBasica.objects.all()

    # Crear archivo ZIP en memoria
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # 1. Crear Excel consolidado con TODO el personal
        wb = Workbook()
        ws = wb.active
        ws.title = "Personal Completo"

        # Título
        ws['A1'] = "REGISTRO COMPLETO DE PERSONAL"
        ws['A1'].font = Font(bold=True, size=14, color="2C3E50")
        ws.merge_cells('A1:L1')

        # Encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ["Cédula", "Nombre Completo", "Género", "Teléfono", "Correo",
                   "Profesión", "Área Conocimiento", "Tipo Perfil", "Experiencia",
                   "Tiempo Experiencia", "Cantidad", "Observaciones"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Datos
        row = 4
        for applicant in applicants:
            ws.cell(row=row, column=1, value=applicant.cedula).border = border
            ws.cell(row=row, column=2, value=applicant.nombre_completo).border = border
            ws.cell(row=row, column=3, value=applicant.genero).border = border
            ws.cell(row=row, column=4, value=applicant.telefono).border = border
            ws.cell(row=row, column=5, value=applicant.correo).border = border
            ws.cell(row=row, column=6, value=applicant.profesion or "N/A").border = border
            ws.cell(row=row, column=7, value=applicant.area_conocimiento or "N/A").border = border
            ws.cell(row=row, column=8, value=applicant.tipo_perfil or "N/A").border = border
            ws.cell(row=row, column=9, value=applicant.experiencia or "N/A").border = border
            ws.cell(row=row, column=10, value=applicant.tiempo_experiencia or "N/A").border = border
            ws.cell(row=row, column=11, value=applicant.cantidad or "N/A").border = border
            ws.cell(row=row, column=12, value=applicant.observacion or "N/A").border = border
            row += 1

        # Ajustar anchos de columna
        for col in range(1, 13):
            ws.column_dimensions[chr(64 + col)].width = 20

        # Guardar Excel consolidado
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        zip_file.writestr("Personal_Completo.xlsx", excel_buffer.getvalue())

        # 2. Agregar Excel individual de cada persona
        for applicant in applicants:
            wb_individual = create_excel_for_person(applicant)
            excel_individual_buffer = io.BytesIO()
            wb_individual.save(excel_individual_buffer)
            excel_individual_buffer.seek(0)

            filename_safe = applicant.nombre_completo.replace(' ', '_')
            zip_file.writestr(
                f"Personal/{filename_safe}/{filename_safe}_Informacion.xlsx",
                excel_individual_buffer.getvalue()
            )

            # 2.1. Agregar PDF ANEXO 11 de cada persona
            try:
                pdf_buffer_individual = generar_anexo11_pdf(applicant)
                zip_file.writestr(
                    f"Personal/{filename_safe}/{filename_safe}_ANEXO_11.pdf",
                    pdf_buffer_individual.getvalue()
                )
            except Exception as e:
                logger.error(f"Error al generar PDF ANEXO 11 para {applicant.nombre_completo}: {str(e)}")

            # 3. Agregar certificados de cada persona
            for idx, experiencia in enumerate(applicant.experiencias_laborales.all(), start=1):
                if experiencia.certificado_laboral:
                    try:
                        certificado_file = experiencia.certificado_laboral
                        with certificado_file.open('rb') as f:
                            file_content = f.read()

                        # Obtener la extensión del archivo
                        # Cloudinary puede no incluir extensión en el nombre, así que intentamos obtenerla de la URL
                        ext = os.path.splitext(certificado_file.name)[1]
                        if not ext and hasattr(certificado_file, 'url'):
                            # Intentar obtener extensión de la URL de Cloudinary
                            url = certificado_file.url
                            # La URL de Cloudinary tiene formato: .../upload/v123456/archivo.ext
                            if '.' in url.split('/')[-1]:
                                ext = '.' + url.split('/')[-1].split('.')[-1].split('?')[0]

                        # Si aún no hay extensión, detectar por contenido
                        if not ext:
                            # Detectar tipo por magic bytes
                            if file_content.startswith(b'%PDF'):
                                ext = '.pdf'
                            elif file_content.startswith(b'\x89PNG'):
                                ext = '.png'
                            elif file_content.startswith(b'\xff\xd8\xff'):
                                ext = '.jpg'
                            else:
                                ext = '.pdf'  # Default a PDF si no se puede detectar

                        cargo_safe = experiencia.cargo.replace(' ', '_').replace('/', '-')

                        zip_file.writestr(
                            f"Personal/{filename_safe}/Certificados/{idx}_{cargo_safe}{ext}",
                            file_content
                        )
                    except Exception as e:
                        logger.error(f"Error al agregar certificado {idx} de {applicant.nombre_completo}: {e}")

    # Preparar respuesta
    zip_buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="Personal_Completo_{timestamp}.zip"'

    return response