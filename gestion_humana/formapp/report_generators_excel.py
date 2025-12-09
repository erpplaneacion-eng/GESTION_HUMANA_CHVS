from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .services import obtener_experiencias_historicas, obtener_resumen_experiencia_historica

def create_excel_for_person(applicant):
    """Crea un archivo Excel con toda la información de una persona"""
    wb = Workbook()

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="2C3E50")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Hoja 1: Información Básica
    ws1 = wb.active
    ws1.title = "Información Básica"

    # Título
    ws1['A1'] = f"INFORMACIÓN PERSONAL - {applicant.nombre_completo}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')

    # Datos personales
    row = 3
    personal_data = [
        ("Cédula", applicant.cedula),
        ("Nombre Completo", applicant.nombre_completo),
        ("Género", applicant.genero),
        ("Dirección", f"{applicant.tipo_via} {applicant.numero_via} #{applicant.numero_casa}"),
        ("Complemento Dirección", applicant.complemento_direccion or "N/A"),
        ("Barrio", applicant.barrio or "N/A"),
        ("Teléfono", applicant.telefono),
        ("Correo", applicant.correo),
    ]

    for label, value in personal_data:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'A{row}'].fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        ws1[f'B{row}'] = value
        ws1[f'A{row}'].border = border
        ws1[f'B{row}'].border = border
        row += 1

    # Información profesional
    row += 2
    ws1[f'A{row}'] = "INFORMACIÓN PROFESIONAL"
    ws1[f'A{row}'].font = title_font
    ws1.merge_cells(f'A{row}:B{row}')
    row += 2

    professional_data = [
        ("Perfil", applicant.perfil or "N/A"),
        ("Área del Conocimiento", applicant.area_del_conocimiento or "N/A"),
        ("Profesión", applicant.profesion or "N/A"),
        ("Contrato", applicant.contrato or "N/A"),
        ("Observaciones", applicant.observacion or "N/A"),
    ]

    for label, value in professional_data:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'A{row}'].fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        ws1[f'B{row}'] = value
        ws1[f'A{row}'].border = border
        ws1[f'B{row}'].border = border
        row += 1

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 50

    # Hoja 2: Experiencia Laboral
    ws2 = wb.create_sheet("Experiencia Laboral")
    ws2['A1'] = f"EXPERIENCIA LABORAL - {applicant.nombre_completo}"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:H1')

    # Encabezados
    headers = ["Cargo", "Cargo Anexo 11", "Fecha Inicial", "Fecha Terminación",
               "Meses", "Días", "Objeto Contractual", "Funciones"]
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row = 4
    for exp in applicant.experiencias_laborales.all():
        ws2.cell(row=row, column=1, value=exp.cargo).border = border
        ws2.cell(row=row, column=2, value=exp.cargo_anexo_11).border = border
        ws2.cell(row=row, column=3, value=exp.fecha_inicial.strftime('%Y-%m-%d')).border = border
        ws2.cell(row=row, column=4, value=exp.fecha_terminacion.strftime('%Y-%m-%d')).border = border
        ws2.cell(row=row, column=5, value=exp.meses_experiencia).border = border
        ws2.cell(row=row, column=6, value=exp.dias_experiencia).border = border
        ws2.cell(row=row, column=7, value=exp.objeto_contractual).border = border
        ws2.cell(row=row, column=8, value=exp.funciones).border = border
        row += 1

    for col in range(1, 9):
        ws2.column_dimensions[chr(64 + col)].width = 20

    # Hoja 2.5: Experiencia Histórica (Base de Datos Caquetá)
    experiencias_historicas = obtener_experiencias_historicas(applicant.cedula)
    if experiencias_historicas:
        resumen_historico = obtener_resumen_experiencia_historica(applicant.cedula)

        ws_hist = wb.create_sheet("Experiencia Histórica")
        ws_hist['A1'] = f"EXPERIENCIA HISTÓRICA (CONTRATOS 2017-2025) - {applicant.nombre_completo}"
        ws_hist['A1'].font = title_font
        ws_hist.merge_cells('A1:H1')

        # Resumen
        ws_hist['A3'] = "Resumen de Experiencia Histórica"
        ws_hist['A3'].font = Font(bold=True, size=11)
        ws_hist['A3'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        ws_hist.merge_cells('A3:B3')

        ws_hist['A4'] = "Total Contratos:"
        ws_hist['B4'] = resumen_historico['total_contratos']
        ws_hist['A5'] = "Experiencia Total:"
        ws_hist['B5'] = resumen_historico['experiencia_texto']

        for row_num in [4, 5]:
            ws_hist[f'A{row_num}'].font = Font(bold=True)
            ws_hist[f'A{row_num}'].border = border
            ws_hist[f'B{row_num}'].border = border

        # Encabezados
        headers_hist = ["N°", "Contrato", "Fecha Inicio", "Fecha Fin",
                       "Días Brutos", "Días Reales", "Traslape", "Observación"]
        for col, header in enumerate(headers_hist, start=1):
            cell = ws_hist.cell(row=7, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Datos
        row = 8
        for idx, exp_hist in enumerate(experiencias_historicas, start=1):
            ws_hist.cell(row=row, column=1, value=idx).border = border
            ws_hist.cell(row=row, column=2, value=exp_hist.contrato).border = border
            ws_hist.cell(row=row, column=3, value=exp_hist.fecha_inicio.strftime('%Y-%m-%d')).border = border
            ws_hist.cell(row=row, column=4, value=exp_hist.fecha_fin.strftime('%Y-%m-%d')).border = border
            ws_hist.cell(row=row, column=5, value=exp_hist.dias_brutos).border = border
            ws_hist.cell(row=row, column=6, value=exp_hist.dias_reales_contribuidos).border = border
            ws_hist.cell(row=row, column=7, value=exp_hist.traslape).border = border
            ws_hist.cell(row=row, column=8, value=exp_hist.explicacion_detallada or "").border = border
            row += 1

        for col in range(1, 9):
            ws_hist.column_dimensions[chr(64 + col)].width = 18

    # Hoja 2.1: Educación Básica (Bachiller)
    ws_basica = wb.create_sheet("Bachiller")
    ws_basica['A1'] = f"EDUCACIÓN BÁSICA (BACHILLER) - {applicant.nombre_completo}"
    ws_basica['A1'].font = title_font
    ws_basica.merge_cells('A1:C1')

    headers_basica = ["Institución", "Año Grado", "Título"]
    for col, header in enumerate(headers_basica, start=1):
        cell = ws_basica.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row = 4
    for basica in applicant.educacion_basica.all():
        ws_basica.cell(row=row, column=1, value=basica.institucion).border = border
        ws_basica.cell(row=row, column=2, value=basica.anio_grado).border = border
        ws_basica.cell(row=row, column=3, value=basica.titulo).border = border
        row += 1
    
    for col in range(1, 4):
        ws_basica.column_dimensions[chr(64 + col)].width = 25

    # Hoja 2.2: Educación Superior (Técnico/Tecnólogo)
    ws_superior = wb.create_sheet("Técnico y Tecnólogo")
    ws_superior['A1'] = f"EDUCACIÓN SUPERIOR (TÉCNICO/TECNÓLOGO) - {applicant.nombre_completo}"
    ws_superior['A1'].font = title_font
    ws_superior.merge_cells('A1:E1')

    headers_superior = ["Nivel", "Institución", "Título", "Fecha Grado", "Tarjeta Profesional"]
    for col, header in enumerate(headers_superior, start=1):
        cell = ws_superior.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row = 4
    for superior in applicant.educacion_superior.all():
        ws_superior.cell(row=row, column=1, value=superior.nivel).border = border
        ws_superior.cell(row=row, column=2, value=superior.institucion).border = border
        ws_superior.cell(row=row, column=3, value=superior.titulo).border = border
        ws_superior.cell(row=row, column=4, value=superior.fecha_grado.strftime('%Y-%m-%d')).border = border
        ws_superior.cell(row=row, column=5, value=superior.tarjeta_profesional or "N/A").border = border
        row += 1
    
    for col in range(1, 6):
        ws_superior.column_dimensions[chr(64 + col)].width = 25

    # Hoja 3: Información Académica
    ws3 = wb.create_sheet("Información Académica")
    ws3['A1'] = f"INFORMACIÓN ACADÉMICA - {applicant.nombre_completo}"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:F1')

    # Encabezados
    headers = ["Profesión", "Universidad", "Tarjeta Profesional",
               "N° Tarjeta/Resolución", "Fecha Expedición", "Fecha Grado"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row = 4
    for academica in applicant.formacion_academica.all():
        ws3.cell(row=row, column=1, value=academica.profesion).border = border
        ws3.cell(row=row, column=2, value=academica.universidad).border = border
        ws3.cell(row=row, column=3, value=academica.tarjeta_profesional).border = border
        ws3.cell(row=row, column=4, value=academica.numero_tarjeta_resolucion or "N/A").border = border
        ws3.cell(row=row, column=5, value=academica.fecha_expedicion.strftime('%Y-%m-%d') if academica.fecha_expedicion else "N/A").border = border
        ws3.cell(row=row, column=6, value=academica.fecha_grado.strftime('%Y-%m-%d')).border = border
        row += 1

    for col in range(1, 7):
        ws3.column_dimensions[chr(64 + col)].width = 20

    # Hoja 4: Posgrados
    ws4 = wb.create_sheet("Posgrados")
    ws4['A1'] = f"POSGRADOS - {applicant.nombre_completo}"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:C1')

    # Encabezados
    headers = ["Nombre Posgrado", "Universidad", "Fecha Terminación"]
    for col, header in enumerate(headers, start=1):
        cell = ws4.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row = 4
    for posgrado in applicant.posgrados.all():
        ws4.cell(row=row, column=1, value=posgrado.nombre_posgrado).border = border
        ws4.cell(row=row, column=2, value=posgrado.universidad).border = border
        ws4.cell(row=row, column=3, value=posgrado.fecha_terminacion.strftime('%Y-%m-%d')).border = border
        row += 1

    for col in range(1, 4):
        ws4.column_dimensions[chr(64 + col)].width = 25

    # Hoja 5: Especializaciones
    ws5 = wb.create_sheet("Especializaciones")
    ws5['A1'] = f"ESPECIALIZACIONES - {applicant.nombre_completo}"
    ws5['A1'].font = title_font
    ws5.merge_cells('A1:C1')

    # Encabezados
    headers = ["Nombre Especialización", "Universidad", "Fecha Terminación"]
    for col, header in enumerate(headers, start=1):
        cell = ws5.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos
    row = 4
    for especializacion in applicant.especializaciones.all():
        ws5.cell(row=row, column=1, value=especializacion.nombre_especializacion).border = border
        ws5.cell(row=row, column=2, value=especializacion.universidad).border = border
        ws5.cell(row=row, column=3, value=especializacion.fecha_terminacion.strftime('%Y-%m-%d')).border = border
        row += 1

    for col in range(1, 4):
        ws5.column_dimensions[chr(64 + col)].width = 25

    # Hoja 6: Cálculo de Experiencia
    ws6 = wb.create_sheet("Cálculo Experiencia")
    ws6['A1'] = f"CÁLCULO DE EXPERIENCIA - {applicant.nombre_completo}"
    ws6['A1'].font = title_font
    ws6.merge_cells('A1:B1')

    row = 3
    try:
        calculo = applicant.calculo_experiencia
        calculo_data = [
            ("Total Meses Experiencia", calculo.total_meses_experiencia),
            ("Total Días Experiencia", calculo.total_dias_experiencia),
            ("Total Experiencia (Años)", calculo.total_experiencia_anos),
            ("Años y Meses", calculo.anos_y_meses_experiencia),
        ]
    except:
        calculo_data = [
            ("Total Meses Experiencia", "No calculado"),
            ("Total Días Experiencia", "No calculado"),
            ("Total Experiencia (Años)", "No calculado"),
            ("Años y Meses", "No calculado"),
        ]

    for label, value in calculo_data:
        ws6[f'A{row}'] = label
        ws6[f'A{row}'].font = Font(bold=True)
        ws6[f'A{row}'].fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        ws6[f'B{row}'] = value
        ws6[f'A{row}'].border = border
        ws6[f'B{row}'].border = border
        row += 1

    ws6.column_dimensions['A'].width = 30
    ws6.column_dimensions['B'].width = 30

    return wb

